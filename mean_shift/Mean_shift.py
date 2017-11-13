import numpy as np
import xlwt
from openpyxl import load_workbook
from sklearn.cluster import MeanShift, estimate_bandwidth #импорт алгоритма
from sklearn.datasets.samples_generator import make_blobs  #импорт библиотеки для генерации

def read_Excel(var_vivoda): #чтение центров
    wb = load_workbook('data.xlsx')
    sheet = wb.get_sheet_by_name('sheetname')

    centers =[[float(sheet['C7'].value),float(sheet['C8'].value)],[float(sheet['D7'].value),float(sheet['D8'].value)],[float(sheet['E7'].value),float(sheet['E8'].value)]]
    n_samples = sheet['E10'].value
    
    print (n_samples)
    
    X, y = make_blobs(n_samples=10000, centers=centers, cluster_std=0.6) #генерация случайных чисел
 
    func(var_vivoda,X,y)
##############################################################################
def read_RANDOM(var_vivoda): #Генерация данных рандомного значения
    # Генерация данных
    centers = [[1, 1], [-1, -1], [1, -1], [3, 3]] #задание количества центров
    X, y = make_blobs(n_samples=10000, centers=centers, cluster_std=0.6) #генерация случайных чисел
 
    func(var_vivoda, X, y) #вызов алгоритма

##############################################################################   
def func(var_vivoda,X,y): #функция вычисляет кластеры. var_vivoda - вариант вывода. График или Excel 
    
    # Compute clustering with MeanShift
    
    # The following bandwidth can be automatically detected using
    bandwidth = estimate_bandwidth(X, quantile=0.2, n_samples=500)
    
    ms = MeanShift(bandwidth=bandwidth, bin_seeding=True)
    ms.fit(X)
    labels = ms.labels_
    cluster_centers = ms.cluster_centers_
    
    labels_unique = np.unique(labels)
    n_clusters_ = len(labels_unique)
    
    print("number of estimated clusters : %d" % n_clusters_)
    print("b=", var_vivoda) #для Debug
    
    #выбор вида дальнейшего отчета
    if var_vivoda==11:
        Plot_result(n_clusters_, labels, X, cluster_centers)
    if var_vivoda==12:
        Excel_result(X,y,cluster_centers)
##############################################################################    
def Excel_result(X,y,cluster_centers):
       
  #  print("DEBUG!!!")
    book = xlwt.Workbook(encoding="utf-8")

    # Добавить лист в книгу 
    sheet1 = book.add_sheet("List_1") 
    
    sheet1.write(2,0, "Координата точки на графике Х*")
    sheet1.write(2,1, "Координата точки на графике Y*")
    sheet1.write(2,4, "Координата центра найденного класстера Х*")
    sheet1.write(2,5, "Координата центра найденного класстера Y*")
    
    for x in range(len(y)):
        sheet1.write(x+3,1,float(X[x,1]))
    #---------------------------------------
    for x in range(len(y)):
        sheet1.write(x+3,0,float(X[x,0]))
    
    #---------------------------------------
    for i in range(len(cluster_centers)):
        for j in range(len(cluster_centers[i])):
            sheet1.write(i+3,j+3,cluster_centers[i,j])
#    print("DEBUG!!!")
     
    book.save("Raport.xls")    
##############################################################################
def Plot_result(n_clusters_, labels, X, cluster_centers):
    # ########################################################################
    # Plot result
    import matplotlib.pyplot as plt
    from itertools import cycle
    
    plt.figure(1)
    plt.clf()
    
    colors = cycle('bgrcmykbgrcmykbgrcmykbgrcmyk')
    for k, col in zip(range(n_clusters_), colors):
        my_members = labels == k
        cluster_center = cluster_centers[k]
        plt.plot(X[my_members, 0], X[my_members, 1], col + '.')
        plt.plot(cluster_center[0], cluster_center[1], 'o', markerfacecolor=col,
                 markeredgecolor='k', markersize=14)
    plt.title('Estimated number of clusters: %d' % n_clusters_)
    plt.show()
##############################################################################
def main():
    print("Input Method:")
    print("1 - Random")
    print("2 - Excel")
    print("3 - Sitet")
    
    print("Input Method =")

    a = int(input())
    #print("Sposob vvoda =",a)
    
    print("-----------------------------------")
    print("Output method")
    print("1 - Plot")
    print("2 - Excel")
    
    print("Output method =")
    
    b = 10 + int(input())
    #print("Sposob vivoda =", b)

    if a==1:
        read_RANDOM(b)
    
    if(a==2):
        read_Excel(b)

if __name__ == "__main__":
	main()