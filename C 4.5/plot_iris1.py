import numpy as np
import numpy as np
import openpyxl
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.datasets import load_iris
from sklearn.datasets.samples_generator import make_blobs  #импорт библиотеки для генерации
from openpyxl import load_workbook
from sklearn.tree import DecisionTreeClassifier
from reportlab.lib.styles import ParagraphStyle


import SimpleExcel

#значения по умолчанию
#cel_0_1 = 0
#cel_1_1 = 1

#cel_2_1 = 0
#cel_2_2 = 2

#cel_3_1 = 0
#cel_3_3 = 3

#cel_4_1 = 1
#cel_4_3 = 2

#cel_5_1 = 1
#cel_5_3 = 3

#cel_6_1 = 2
#cel_6_3 = 3

#для определения цвета
'''def getBGColor(book, sheet, row, col):
    xfx = sheet.cell_xf_index(row, col)
    xf = book.xf_list[xfx]
    bgx = xf.background.pattern_colour_index
    pattern_colour = book.colour_map[bgx]
    return pattern_colour
'''
def read_RANDOM(var_vivoda, var_vvoda): #только для этого метода, т.к. странная генерация данных 
    # Parameters1
   # mass = [[0,0]]
    
    n_classes = 3
    plot_colors = "bry"
    plot_step = 0.02
    mass = [[0, 1], [0, 2], [0, 3], [1, 2], [1, 3], [2, 3]] #стандартные параметры
       
    func(var_vivoda, n_classes, plot_colors, plot_step, var_vvoda, mass)

def read_Excel(var_vivoda, var_vvoda):
    
    var_vvoda = 1 #переопределить на вариант ввода RANDOM
    
    wb = xlrd.open_workbook('data.xls',formatting_info=True)
    sheet = wb.sheet_by_index(0) 
    
    row = 0
    col = 0    

    plot_colors = getBGColor(wb, sheet, row, col)  #получение цвета ячейки. 
    
    print('Color=',plot_colors)
  
    cell_tmp = sheet.cell(row, col)
    
    n_samples = valueInt(cell_tmp)    ###########?????????????????????????
    print('n_samples=',n_samples)
    #if cell_tmp.ctype==xlrd.XL_CELL_NUMBER:
    #    n_samples = cell_tmp.value
    #    print('n_samples=',n_samples)
        
    read_RANDOM(var_vivoda, var_vvoda)
    
    
def func(var_vivoda, n_classes, plot_colors, plot_step, var_vvoda, mass):
    
    # Load data
    iris = load_iris()   #подгрузка бд из интернета, судя по документации
        
    X =[0,0]
    y = [0,0]
    #здесь так же подставляются значения из Excel. Поэтому такие разные имена ячеек
    for pairidx, pair in enumerate(mass):
    
        if var_vvoda==1:
            # автоматическая генерация значений
            X = iris.data[:, pair]
            y = iris.target
        
        # Train
        clf = DecisionTreeClassifier().fit(X, y)
    
        # Plot the decision boundary
        plt.subplot(2, 3, pairidx + 1)
    
        x_min, x_max = X[:, 0].min() - 1, X[:, 0].max() + 1
        y_min, y_max = X[:, 1].min() - 1, X[:, 1].max() + 1
        xx, yy = np.meshgrid(np.arange(x_min, x_max, plot_step),
                             np.arange(y_min, y_max, plot_step))
    
        Z = clf.predict(np.c_[xx.ravel(), yy.ravel()])
        Z = Z.reshape(xx.shape)
        cs = plt.contourf(xx, yy, Z, cmap=plt.cm.Paired)
    
        plt.xlabel(iris.feature_names[pair[0]])
        plt.ylabel(iris.feature_names[pair[1]])
        plt.axis("tight")
    
        # Plot the training points
        for i, color in zip(range(n_classes), plot_colors):
            idx = np.where(y == i)
            plt.scatter(X[idx, 0], X[idx, 1], c=color, label=iris.target_names[i],
                        cmap=plt.cm.Paired)
 
        
#######################-Вывод результатов в Excel-###########################           
    if var_vivoda == 12: #если установлен вывод в Excel
        Excel_result(X,y)
#############################################################################                        
            
    plt.axis("tight")
    
    plt.suptitle("Decision surface of a decision tree using paired features")
    plt.legend()
    plt.show()
#############################################################################    

def Excel_result(X,y):
       
    print("DEBUG!!!")
    book = xlwt.Workbook(encoding="utf-8")
    
    sheet1 = book.add_sheet("List_1") 
    
    for x in range(len(X)):
        sheet1.write(x+3,0,float(X[x,0]))
                        
    for x in range(len(X)):
        sheet1.write(x+3,1,float(X[x,1]))
            
    for x in range(len(y)):
        sheet1.write(x+3,2,float(y[x]))          
     
    book.save("Raport.xls")
  

############################################################################
############################################################################
def main():
    print("Sposob vvoda")
    print("1 - Random")
    print("2 - Excel")
    print("3 - sait")
    
    print("Sposob vvoda =")

    a = int(input())
    print("Sposob vvoda =",a)
    
    print("-----------------------------------")
    print("Sposob vivoda")
    print("1 - graphic")
    print("2 - Excel")
    
    print("Sposob vivoda =")
    
    print("-----------------------------------")
    b = 10 + int(input())
    print("Sposob vvoda =", a)
    print("Sposob vivoda =", b)
    print("-----------------------------------")


    if a==1:
        read_RANDOM(b,a)
        print("1!")
    
    if a==2:
        read_Excel(b,a)
       # print("2")

if __name__ == "__main__":
	main()