import numpy as np
import numpy
import openpyxl
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.datasets import load_iris
from sklearn.datasets.samples_generator import make_blobs  #импорт библиотеки для генерации
from openpyxl import load_workbook
from sklearn.tree import DecisionTreeClassifier
from reportlab.lib.styles import ParagraphStyle
import random



mass = [[0, 1], [0, 2], [0, 3], [1, 2], [1, 3], [2, 3]] #стандартные параметры

def read_RANDOM(var_vivoda, var_vvoda): #только для этого метода, т.к. странная генерация данных 
    
    n_classes = 3
    plot_colors = "bry"
    plot_step = 0.02

    func(var_vivoda, n_classes, plot_colors, plot_step, var_vvoda)

def read_Excel(var_vivoda, var_vvoda):
    
  #  var_vvoda = 1 #переопределить на вариант ввода RANDOM
    
    wb = xlrd.open_workbook('data.xls',formatting_info=True)
    sheet = wb.sheet_by_index(0) 
    
    row = 0
    col = 0    
 
    cell_tmp = sheet.cell(row, col)
    
    n_samples = sheet.cell(row, col) #valueInt(cell_tmp)    ###########?????????????????????????
    print('n_samples=',n_samples)
    
    n_classes = 3 #значение по умолчанию
    plot_colors = "bry" #значение по умолчанию
    plot_step = 0.02 #значения по умолчанию 
    
    func(var_vivoda, n_classes, plot_colors, plot_step, var_vvoda)
    
def read_iris(var_vivoda, var_vvoda):
    n_classes = 3 #значение по умолчанию
    plot_colors = "bry" #значение по умолчанию
    plot_step = 0.02 #значения по умолчанию 
    func(var_vivoda, n_classes, plot_colors, plot_step, var_vvoda)
    

def func(var_vivoda, n_classes, plot_colors, plot_step, var_vvoda):
    
    # Load data
   # iris = load_iris()   #подгрузка бд из интернета, судя по документации
        
   
    #здесь так же подставляются значения из Excel. Поэтому такие разные имена ячеек
    for pairidx, pair in enumerate(mass):
        iris = load_iris()
        
        if var_vvoda==1:  #генеарция через Random
            #iris = load_iris() #посмотреть на сгенеренные значения
            # автоматическая генерация значений
            X = iris.data[:, pair]
          #  print('x=', X)
            y = iris.target
            print('len_X=', len(X))
           
        if var_vvoda==3:  #генеарция через Iris
         #   iris = load_iris() #посмотреть на сгенеренные значения
            # автоматическая генерация значений
            X = iris.data[:, pair]
            y = iris.target
        
        if var_vvoda==2:
            filename = 'diabetes.csv'
            raw_data = open(filename, 'rt')
            X = numpy.loadtxt(raw_data, delimiter=",")
            y=iris.target
            
        # Train
        clf = DecisionTreeClassifier().fit(X, y)
    
        # Plot the decision boundary
        plt.subplot(2, 3, pairidx + 1)
    
        x_min, x_max = X[:, 0].min() - 1, X[:, 0].max() + 1
        y_min, y_max = X[:, 1].min() - 1, X[:, 1].max() + 1
        xx, yy = np.meshgrid(np.arange(x_min, x_max, plot_step),
                             np.arange(y_min, y_max, plot_step))
    
        Z = clf.predict(np.c_[xx.ravel(), yy.ravel()])
        Z = Z.reshape(xx.shape)
        cs = plt.contourf(xx, yy, Z, cmap=plt.cm.Paired)
    
        plt.xlabel(iris.feature_names[pair[0]])
        plt.ylabel(iris.feature_names[pair[1]])
        plt.axis("tight")
    
        # Plot the training points
        for i, color in zip(range(n_classes), plot_colors):
            idx = np.where(y == i)
            plt.scatter(X[idx, 0], X[idx, 1], c=color, label=iris.target_names[i],
                        cmap=plt.cm.Paired)
 
        
#######################-Вывод результатов в Excel-###########################           
    if var_vivoda == 12: #если установлен вывод в Excel
        Excel_result(X,y)
#############################################################################                        
            
    plt.axis("tight")
    
    plt.suptitle("Decision surface of a decision tree using paired features")
    plt.legend()
    plt.show()
#############################################################################    

def Excel_result(X,y):
       
    print("DEBUG!!!")
    book = xlwt.Workbook(encoding="utf-8")
    
    sheet1 = book.add_sheet("List_1") 
    
    for x in range(len(X)):
        sheet1.write(x+3,0,float(X[x,0]))
                        
    for x in range(len(X)):
        sheet1.write(x+3,1,float(X[x,1]))
            
    for x in range(len(y)):
        sheet1.write(x+3,2,float(y[x]))          
     
    book.save("Raport.xls")
  

############################################################################
############################################################################
def main():


    print("Input Method:")
    print("1 - Random")
    print("2 - Excel")
    print("3 - Sitet")
    
    print("Input Method =")

    a = int(input())
#    print("Sposob vvoda =",a)
    
    print("-----------------------------------")
    print("Output method")
    print("1 - Plot")
    print("2 - Excel")
    
    print("Output method =")
    
 #   print("-----------------------------------")
    b = 10 + int(input())

    if a==1:
        read_RANDOM(b,a)
      #  print("1!")
    
    if a==2:
        read_Excel(b,a)
       # print("2")
       
    if a==3: # read iris_database
        read_iris(b,a)
       # print("2")   

if __name__ == "__main__":
	main()