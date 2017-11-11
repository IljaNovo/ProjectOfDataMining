# -*- coding: utf-8 -*-
"""
Created on Tue Oct 31 14:05:07 2017

EN:
    
    library for simplified work with Excel.

    Contains the following methods:
        -obtaining the color of the cell;
        -getting number from cell in Excel
        -
        -
    
RU:    
    Библеотека для упрощенной работы с Excel
    
    Содердит следующие методы:
        -получение цвета ячейки;
        -получение числа из ячейки в Excel
        -
        -
        -
@author: SIO
"""
import numpy as np
import numpy as np
import openpyxl
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.datasets import load_iris
from sklearn.datasets.samples_generator import make_blobs  #импорт библиотеки для генерации
from openpyxl import load_workbook
from sklearn.tree import DecisionTreeClassifier
from reportlab.lib.styles import ParagraphStyle

'''Определение цвета ячейки. Возвращает цвет в формате RGB
Входные данные:
    -book - книга. Инициализированная книга
    -sheet - открытый в данный момент лист Excel
    -row - столбец. Исчисление от 0
    -col - строка. Исчисление от 0
'''
def getBGColor(book, sheet, row, col):
    xfx = sheet.cell_xf_index(row, col)
    xf = book.xf_list[xfx]
    bgx = xf.background.pattern_colour_index
    pattern_colour = book.colour_map[bgx]
    return pattern_colour

#Получение числа из ячейки в Excel
def valueInt(cell_tmp):
    if cell_tmp.ctype==xlrd.XL_CELL_NUMBER:
        n_samples = cell_tmp.value
        return n_samples
    else:
        return 'false' ##############################????????????????
    
    
        